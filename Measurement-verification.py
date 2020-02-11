'''
Takes torque measurements and saves it directly to an excel file in Automated-Testing folder with the torque setpoint and description as the file name.
'''

import os
import time
import sys
import math as mt
import numpy as np
import subprocess
import json

import pigpio as io # pigpio daemon
from pipyadc import ADS1256 #Library for interfacing with the ADC via Python
import gpiozero as gz #Library for using the GPIO with python
import pandas as pd
import RPi.GPIO as GPIO # Using GPIO instead of wiringpi for reading digital inputs
from openpyxl import Workbook

import Tconfigs as tcf 
import gcpConfigs as gcpc
from ADS1256_definitions import * #Configuration file for the ADC settings
import dac8552.dac8552 as dac1
from dac8552.dac8552 import DAC8552, DAC_A, DAC_B, MODE_POWER_DOWN_100K #Library for using the DAC

# Set pin numbers for the relay channels and the limit switch inputs
# Note that the pin numbers here follow the wiringPI scheme, which we've setup for *.phys or the GPIO header locations
# Since the wiringpi module communicates through the GPIO, there shouldn't be a need to initiate the SPI bus connection

chan = ('1', '2')
test = []
i = 0
nos = 0
yes = ('YES', 'yes', 'y', 'Ye', 'Y')
no = ('NO','no', 'n', 'N', 'n')
yes_no = ('YES', 'yes', 'y', 'Ye', 'Y', 'NO','no', 'n', 'N', 'n')

# DAC/ADS setup
ads = ADS1256()
ads.cal_self() 
astep = ads.v_per_digit
dac = DAC8552()
dac.v_ref = 5
step = dac.digit_per_v

######################## Original Code and Function Definitions from the pipyadc library ################################################
EXT1, EXT2, EXT3, EXT4 = POS_AIN0|NEG_AINCOM, POS_AIN1|NEG_AINCOM, POS_AIN2|NEG_AINCOM, POS_AIN3|NEG_AINCOM
EXT5, EXT6, EXT7, EXT8 = POS_AIN4|NEG_AINCOM, POS_AIN5|NEG_AINCOM, POS_AIN6|NEG_AINCOM, POS_AIN7|NEG_AINCOM

INPUTS_ADDRESS = (EXT1, EXT2, EXT3, EXT4, EXT5, EXT6, EXT7, EXT8)

CH1_Loc = {'cntrl' : DAC_A,
           'torq' : INPUTS_ADDRESS[0],
           'FK_On': 6,
           'FK_Off' : 13} #GPIO pin numbers

CH2_Loc = {'cntrl' : DAC_B,
           'torq' : INPUTS_ADDRESS[3],
           'FK_On': 19,
           'FK_Off' : 26} #GPIO pin numbers

CH_Out = {'1' : DAC_A ,
          '2' : DAC_B}

tests = ('1', '2')
test = []

test_channels = {0: CH1_Loc,
                 1: CH2_Loc}

binary = {'INPUT' : 0,
          'OUTPUT': 1,
          'LOW' : 0,
          'HIGH' : 1}

OUTPUT = binary['OUTPUT']
INPUT = binary['INPUT']
LOW = binary['LOW']
HIGH = binary['HIGH']

# Set up arrays
vData = []
vtime = []

state = True
tstart = time.time()

#create xl workbook
wb = Workbook()
sheet = wb.active
headers = [('Time (s)', 'Voltage (V)', 'Torque (in-lbs)', 'Average Voltage (V)', 'Average Torque (in-lbs)', 'Constant torque')]
for row in headers:
    sheet.append(row)

def torqueMeasurement(input, cyclepoint):
    # Collect 10 data point readings across 1 second
    y = test[0].pv
    print('pv = ', y)
    setData=[] #array for average torque calculation
    h = 25
    i = 0
    for i in range (0, h):
        raw_channels = ads.read_oneshot(input)
        vo = float(raw_channels*astep) # Convert raw value to voltage
        ti = time.time()-tstart
        # append data
        setData.append(vo)
        print(vo)
        sheet.cell(row=i+2+h*cyclepoint+(h*cyclepoint)*(y-1), column =1).value = ti
        sheet.cell(row=i+2+h*cyclepoint+(h*cyclepoint)*(y-1), column=2).value = vo
        torq = torqueConvert(vo) # Convert voltage value to torque value
        sheet.cell(row=i+2+h*cyclepoint+(h*cyclepoint)*(y-1), column=3).value = torq
        time.sleep(0.05)
    # Remove 6 max and min values
    for x in range (0,6):
        setData.remove(max(setData)) 
        setData.remove(min(setData))
    voltage = float(sum(setData)/len(setData)) # Average everything else
    print('voltage reading: ', voltage) # for troubleshooting/calibration
    sheet.cell(row=11+h*cyclepoint+(h*cyclepoint)*(y-1), column = 4).value = voltage
    to = torqueConvert(voltage) # Convert voltage value to torque value
    sheet.cell(row=11+h*cyclepoint+(h*cyclepoint)*(y-1), column=5).value = to
    print('torque reading:', to)
    return to

def torqueConvert(volt):
    torqueVal = (volt - 2.5)*6000/2.5 # Convert reading to torque value in in-lbs
    return(torqueVal)

def noSwitchCheck(test, testIndex): # Use this switchCheck if not hooked up to limit switches
    '''
    Read the state of the actuator limit switch inputs
    If they changed, do some stuff, if they haven't changed, then do nothing'''

    if test.active == True:
        if (test.pv < test.target): # Check to see if the current cycle count is less than the target
            # collect "cycle_points" amount of points in cycle
            print('test.pv: ', test.pv)
            w = 0
            while w > (test.cycle_points):
                print('w = ', w)
                torqueMeasurement(test_channels[testIndex]['torq'], w)
                time.sleep(1)
                break
            test.pv+= 1 # Increment the pv counter if the switch changed
            w +=1
        else:
            test.active = False
    else:
        pass

def switchCheck(test, testIndex): # Use this switchCheck if hooked up to actuator with limit switches
    '''
    Read the state of the actuator limit switch inputs
    If they changed, do some stuff, if they haven't changed, then do nothing'''

    if test.active == True:
        if (test.pv < test.target): # Check to see if the current cycle count is less than the target
            open_switch = test_channels[testIndex]['FK_On']
            closed_switch = test_channels[testIndex]['FK_Off']
            open_state = GPIO.input(open_switch)
            closed_state = GPIO.input(closed_switch)
            open_last_state = test.open_last_state # Store the last FK_On switch state in a temp variable
            closed_last_state = test.closed_last_state # Store the last FK_Off switch state in temp variable
            
            if (open_last_state == LOW) & (open_state == HIGH) & (closed_state == HIGH): # Check if changed from fully open position to closing (moving)
                test.open_last_state = HIGH # Reset the "open last state" of the switch
                length = time.time() - test.cycle_start - test.cycle_time # Calculate the length of the last duty cycle
                print('pv', test.pv, 'target', test.target)

                if (length > (test.duty_cycle*.5)):
                    print("Switch {} confirmed. Actuator is closing.".format(test.name))
                    test.cycle_start = time.time() # update cycle start time
                    print('cycle start updated to: ', test.cycle_start) # debugging
                    test.pv+= 1 # Increment the pv counter if the switch changed
                    print('test.pv: ', test.pv)

                    # collect "cycle_points" amount of points in cycle
                    w = 0
                    for w in range (test.cycle_points):
                        while True:
                            # wait 1/3 of cycle time or 1/cyclepoints
                            if (time.time() - test.cycle_start) > (((w)/test.cycle_points)*test.cycle_time):
                                torqueMeasurement(test_channels[testIndex]['torq'], w)
                                break
                else:
                    test.bounces = test.bounces + 1
                    print("Switch {} bounced. Bounce count: {}.".format(testIndex, test.bounces))

            elif (closed_last_state == LOW) & (closed_state == HIGH) & (open_state == HIGH): # Check if changed from fully closed position to opening (moving)
                test.closed_last_state = HIGH # Reset the "closed last state" of the switch
                length = time.time() - test.cycle_time # Calculate the length of the last duty cycle
                print('pv', test.pv, 'target', test.target)

                if (length > (test.duty_cycle*.5)):
                    print("Switch {} confirmed. Actuator is opening.".format(test.name))
                    test.cycle_start = time.time() # Update cycle start time
                    test.pv+= 1 # Increment the pv counter if the switch changed
                    print('test.pv: ', test.pv)

                    # collect "cycle_points" amount of points in cycle
                    w=0
                    for w in range (test.cycle_points):
                        # wait 1/3 of cycle time or 1/cyclepoints
                        while True:
                            if (time.time() - test.cycle_start) > (((w)/test.cycle_points)*test.cycle_time):
                                torqueMeasurement(test_channels[testIndex]['torq'], w)
                                break
                else:
                    test.bounces = test.bounces + 1
                    print("Switch {} bounced. Bounce count: {}.".format(testIndex, test.bounces))

            elif (open_last_state == HIGH) & (open_state == LOW) & (closed_state == HIGH): # Check to see if recently in fully open position
                print("Switch {} changed. Actuator is in fully open position.".format(testIndex))
                test.open_last_state = LOW # Update last switch state
                #print('test.cycle_time updated to: ', test.cycle_time)

            elif (closed_last_state == HIGH) & (closed_state == LOW) & (open_state == HIGH): # Check to see if recently in fully closed position
                print("Switch {} changed. Actuator is in fully closed position.".format(testIndex))
                test.closed_last_state = LOW # Update last switch state
                #print('test.cycle_time updated to: ', test.cycle_time)

            else:
                pass
        else:
            test.active = False
    else:
        pass

# prevents issues with shutdown (cogging etc)
def shut_down():
    tcf.running_off() # Turn off test running LED
    i = 0
    for i, in enumerate(test):
        dac.power_down(test[0], 0)
    tcf.killDaemons()

print('Starting test set-up')

while True:
    if (i >= len(chan)): # exit the loop if the test channels are full
        break
    
    print('Add new test on {}? (yes/no) '.format(chan[i]))
    prompt = input() # prompt the user to see if they want to add a new test

    if prompt not in yes_no: # If the input isn't recognized, try again
        print('Input error, please enter yes or no ')
        tcf.warning_on()

    elif prompt in no: # If they enter no, exit the loop
        tcf.warning_off()
        i += 1
        nos += 1

    elif prompt in yes: # If they answer yes, run the test creation functions
        tcf.warning_off()
        test.append(gcpc.define_test()) # Creates a new gcp test class
        test[i-nos].create_on_off_test() # Loads the test parameters

        print('1.taking initial measurements at 0')
        for s in range(0,10):
            raw_channels = ads.read_oneshot(INPUTS_ADDRESS[0])
            vo = float(raw_channels*astep) # Convert raw value to voltage
            print(vo)
            time.sleep(0.1)

        test[i-nos].parameter_check() # Checks that the parameters are within normal working ranges
        tcf.set_on_off(test[i-nos], (i + nos)) # Sets up the IO pins to work for torque tests

        tcf.brakeOn(test[i-nos], (i-nos)) # Turn brake on to setpoint value
        i += 1 # Increment the test channel counter to track the number of active tests

    else:
        tcf.warning_on()
        shut_down()
        raise Warning('Something went wrong :(') # If the test case isn't caught by the above, something's wrong

wait = 0.5 # A small waiting period is necessary, otherwise the switch input reads each cycle multiple times
print('Running test(s)')
tcf.running_on() # Turn on test running LED
stamp = time.time()


'''############################################# Use if hooked up to limit switches ############################################################

while True: # Start a loop to run the torque tests
    # Loop through each test class one by one
    i = 0
    if ((time.time() - stamp) < (wait)): # Check to see if it's time to check the switch inputs again
        pass

    elif test[i].active != True: # Check to see if the test is still active
        pass

    else: 
        switchCheck(test[i], i) # Run a check of the current switch state, add 1 to pv if valid
        stamp = time.time()
        if test[i].pv >= test[i].target:
            state = False
        # Loop through each test class and see if they're all inactive
        state = (state | test[i].active)

    if state == False: # If all the test states are inactive, exit the loop
        break
        
    else:
        pass
else:
    pass

##################################################### end code block for limit switches ####################################################'''


############################################ Use if pi not hooked up to limit switches #####################################################

while True: # Start a loop to run the torque tests
    # Loop through each test class one by one
    i = 0

    if test[i].active != True: # Check to see if the test is still active
        pass

    else: 
        t = 0
        while t < 7:
            print('wait ', t)
            t+=1
            time.sleep(1)
        noSwitchCheck(test[i], i)
        stamp = time.time()
        if test[i].pv >= test[i].target:
            state = False
        # Loop through each test class and see if they're all inactive
        state = (state | test[i].active)

    if state == False: # If all the test states are inactive, exit the loop
        break
        
    else:
        pass
else:
    pass

########################################### end of code block for no limit switches #######################################################


#save excel sheet
wb.save('{} in-lbs_{}.xlsx'.format(test[0].control, test[0].description))
print('excel file saved')

print('about to power down') # in case cycle time was set too short
time.sleep(5)

'''##################Use with limit switches######################
print('powering down')
shut_down()
############################################################'''

############use without limit switches###########################
tcf.killDaemons()
#############################################################

print("Test exited with a clean status")
